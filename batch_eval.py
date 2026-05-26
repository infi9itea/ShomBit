#!/usr/bin/env python3
"""
batch_eval.py

Runs every query in the evaluation Excel sheet through the current
RAG+LLM pipeline (Llama-3.3-70b via Groq) and produces:
  1. <out_xlsx>  — original workbook + new "RAG+LLM (Llama-3.3-70b)" column
  2. <out_jsonl> — one JSON record per query for evaluation.py

Usage (on Kaggle / terminal):
    python batch_eval.py \
        --xlsx  "/kaggle/input/.../ewu_admission_bot_query_sheet.xlsx" \
        --data-dir "/kaggle/input/.../data-1"
"""
import argparse
import json
import logging
import os
import sys
import time

import openpyxl

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# ── Relevant source per intent (1–20, same for all three language variants) ───
# These match the "source:" frontmatter field in your KB markdown files,
# or the scraped URL. Used for retrieval metrics (Hit@k, MRR, nDCG).
RELEVANT_SOURCES: dict[int, list[str]] = {
    1:  ["https://ewubd.edu/undergraduate-dates-deadline",
         "https://ewubd.edu/graduate-dates-deadline"],       # application deadline
    2:  ["EWU Admission JSON"],                              # how to apply
    3:  ["EWU Admission JSON",
         "EWU Admission Requirements JSON"],                 # required documents
    4:  ["EWU Academic Structure JSON"],                     # CSE courses
    5:  ["EWU Academic Structure JSON", "EWU Faculty JSON"], # departments list
    6:  ["EWU Tuition Fees JSON", "EWU Programs JSON"],      # BBA total credits
    7:  ["EWU Institutional Information JSON"],              # why EWU
    8:  ["EWU Tuition Fees JSON"],                           # CSE tuition fee
    9:  ["EWU Scholarships JSON"],                           # scholarship opportunities
    10: ["EWU Scholarships JSON"],                           # merit scholarship terms
    11: ["https://ewubd.edu/undergraduate-tuition-fees",
         "EWU Tuition Fees JSON"],                           # fee change
    12: ["EWU Institutional Information JSON"],              # location
    13: ["EWU Institutional Information JSON"],              # notable alumni
    14: ["EWU Campus Life Facilities JSON",
         "EWU Facilities JSON"],                             # medical center
    15: ["EWU Helpdesk JSON"],                               # registrar contact
    16: ["EWU Helpdesk JSON"],                               # CSE helpdesk email
    17: ["https://ewubd.edu/events"],                        # events this month
    18: ["EWU Admission Requirements JSON"],                 # international students
    19: ["EWU Campus Life Facilities JSON"],                 # hostel/dormitory
    20: ["East West University Student Conduct and Discipline Policy",
         "EWU Programs JSON"],                               # credit transfer
}

# Groq free-tier allows ~30 req/min for Llama-3.3-70b.
# 2 s between queries keeps us safely under the limit.
_INTER_QUERY_DELAY = 2.0
_MAX_RETRIES = 3


def _intent_idx(seq: int) -> int:
    """Maps seq 1-60 to intent 1-20 (same question in three languages)."""
    return ((int(seq) - 1) % 20) + 1


def load_pipeline(data_dir: str):
    """Initialise the full RAG pipeline — mirrors app.py startup."""
    from huggingface_hub import login
    from config import HF_TOKEN
    from data_pipeline import load_markdown_docs, chunk_documents, scrape_dynamic_docs
    from rag_core import build_vectorstore, Reranker, load_llm, build_rag_chain

    if not HF_TOKEN:
        raise ValueError("HF_TOKEN is not set — export it before running.")
    login(HF_TOKEN)

    log.info("Loading documents from %s …", data_dir)
    docs = (
        chunk_documents(load_markdown_docs(data_dir))
        + chunk_documents(scrape_dynamic_docs())
    )
    log.info("Total chunks: %d", len(docs))

    log.info("Building vectorstore …")
    vectorstore, bm25, docs = build_vectorstore(docs)

    log.info("Loading reranker and LLM …")
    reranker = Reranker()
    llm = load_llm()

    chain = build_rag_chain(llm, vectorstore, bm25, docs, reranker)
    log.info("Pipeline ready.")
    return chain


def read_queries(xlsx_path: str) -> list[dict]:
    """Parse query rows from the workbook. Expects columns A-J as in the original sheet."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    queries = []
    for row in rows[1:]:
        if not row[3]:          # skip empty rows (column D = Query)
            continue
        seq          = row[0]
        lang         = row[2] or "unknown"
        question     = row[3]
        ground_truth = row[9] or ""  # column J = Ground Truth
        queries.append({
            "seq":          seq,
            "intent_idx":   _intent_idx(seq),
            "language":     lang,
            "question":     question,
            "ground_truth": ground_truth,
        })

    log.info("Found %d queries (%s)", len(queries),
             ", ".join(f"{l}:{sum(1 for q in queries if q['language']==l)}"
                       for l in ["English", "Bangla", "Banglish"]))
    return queries


def run_query(chain, question: str) -> tuple[str, list, list, float]:
    """Run one query with retry on transient errors. Returns (answer, sources, contexts, latency_s)."""
    for attempt in range(1, _MAX_RETRIES + 1):
        try:
            t0 = time.perf_counter()
            res = chain(question)
            latency = round(time.perf_counter() - t0, 3)
            return res["answer"], res["sources"], res["contexts"], latency
        except Exception as exc:
            log.warning("Attempt %d failed for %r: %s", attempt, question[:40], exc)
            if attempt < _MAX_RETRIES:
                time.sleep(attempt * 5)   # back off before retry
    return f"ERROR after {_MAX_RETRIES} attempts", [], [], -1.0


def run_all(xlsx_path: str, out_xlsx: str, out_jsonl: str, data_dir: str):
    chain   = load_pipeline(data_dir)
    queries = read_queries(xlsx_path)

    results = []
    with open(out_jsonl, "w", encoding="utf-8") as jf:
        for i, q in enumerate(queries, 1):
            log.info("[%d/%d] [%s] %s", i, len(queries), q["language"], q["question"][:70])

            answer, sources, contexts, latency = run_query(chain, q["question"])

            record = {
                "seq":              q["seq"],
                "question":         q["question"],
                "language":         q["language"],
                "answer":           answer,
                "sources":          sources,
                "contexts":         contexts,
                "latency_s":        latency,
                "ground_truth":     q["ground_truth"],
                "relevant_sources": RELEVANT_SOURCES.get(q["intent_idx"], []),
            }
            results.append(record)

            # write incrementally so Kaggle timeouts don't lose progress
            jf.write(json.dumps(record, ensure_ascii=False) + "\n")
            jf.flush()

            log.info("  latency=%.2fs | answer=%s", latency,
                     answer[:80].replace("\n", " "))

            if i < len(queries):
                time.sleep(_INTER_QUERY_DELAY)  # stay within Groq rate limit

    # ── Write new column into Excel ───────────────────────────────
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = "RAG+LLM (Llama-3.3-70b) Response"

    result_by_question = {r["question"]: r for r in results}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        question = row[3]
        if question and question in result_by_question:
            r = result_by_question[question]
            cell_text = r["answer"]
            if r["sources"]:
                cell_text += "\nSources: " + ", ".join(r["sources"])
            cell_text += f"\n[latency: {r['latency_s']}s]"
            ws.cell(row=row_idx, column=new_col).value = cell_text

    wb.save(out_xlsx)
    log.info("Saved updated workbook → %s", out_xlsx)
    log.info("Saved JSONL            → %s", out_jsonl)

    # ── Print a quick summary ─────────────────────────────────────
    by_lang: dict[str, list[float]] = {}
    for r in results:
        lang = r["language"]
        if r["latency_s"] >= 0:
            by_lang.setdefault(lang, []).append(r["latency_s"])
    print("\n=== Latency summary (seconds) ===")
    for lang, lats in sorted(by_lang.items()):
        print(f"  {lang:<10} n={len(lats):>2}  mean={sum(lats)/len(lats):.2f}  "
              f"min={min(lats):.2f}  max={max(lats):.2f}")
    print(f"\nRun  python evaluation.py --gold {out_jsonl} --out results.csv")


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Batch-evaluate all 60 queries.")
    ap.add_argument("--xlsx",      required=True,
                    help="Path to the evaluation Excel file")
    ap.add_argument("--out-xlsx",  default="eval_with_llm.xlsx",
                    help="Where to save the updated workbook (default: eval_with_llm.xlsx)")
    ap.add_argument("--out-jsonl", default="gold.jsonl",
                    help="Where to save the JSONL gold set (default: gold.jsonl)")
    ap.add_argument("--data-dir",
                    default=os.environ.get("DATA_DIR", "./data"),
                    help="KB data directory (default: $DATA_DIR or ./data)")
    args = ap.parse_args()

    run_all(args.xlsx, args.out_xlsx, args.out_jsonl, args.data_dir)
